import os
import openpyxl
current_dir=os.path.dirname(__file__)
fname=os.path.join(current_dir,"assign.xlsx")
wb= openpyxl.Workbook()
sheet=wb.active
alpha=[]
for i in range(65,91):
    alpha.append(chr(i))
sheet.title='Assignment1'
print("Sheet Title:",sheet.title)
print("Sheets:",wb.sheetnames)
'''
wb.create_sheet(index=1,title='try')
sheet=wb['try']
'''
for i in range(1,101):
    for j in alpha:
        sheet[j+ str(i)]=j+str(i)
#print(type(sheet))
wb.save(fname)